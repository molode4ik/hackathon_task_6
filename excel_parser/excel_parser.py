from openpyxl import load_workbook


def read_xlsx(path: str):
    workbook = load_workbook(path)
    return workbook[workbook.sheetnames[0]]

    
def search_start_row(worksheet, row_id):
    for row in worksheet.iter_rows(row_id):
        for cell in row:
            if cell.value == "Местоположение":
                return row_id
            return search_start_row(worksheet, row_id-1)


def create_values_list(worksheet, min_row, max_row):
    data = []
    for row_id, row in enumerate(worksheet.iter_rows(min_row= min_row, max_row= max_row)):
        if row_id == 0 :
            keys = [cell.value for cell in row]
        else:
            data.append({
                key: cell.value for key, cell in zip(keys, row)
                })
    return data


def parse_excel(path):
    worksheet = read_xlsx(path)
    max_row = worksheet.max_row
    min_row = search_start_row(worksheet, max_row)
    data = create_values_list(worksheet, min_row, max_row)
    return data
